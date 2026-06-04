"""
Finance Bot v4 — Multi-user with Google Sheets + Finn AI + Scheduler
"""

import os, re, csv, io, logging, asyncio
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    XLSX_OK = True
except ImportError:
    XLSX_OK = False
from datetime import datetime, timedelta
from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup, WebAppInfo
from telegram.ext import Application, CommandHandler, MessageHandler, CallbackQueryHandler, ContextTypes, filters
import pytz

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

TELEGRAM_TOKEN = os.getenv("TELEGRAM_TOKEN", "")
WEBAPP_URL = os.getenv("WEBAPP_URL", "")
SPREADSHEET_NAME = os.getenv("SPREADSHEET_NAME", "Finance Bot Data")
GEMINI_API_KEY = os.getenv("GEMINI_API_KEY", "")
TIMEZONE = pytz.timezone(os.getenv("BOT_TIMEZONE", "Europe/Berlin"))
DEFAULT_REMINDER_HOUR = int(os.getenv("REMINDER_HOUR", "20"))
DEFAULT_REMINDER_MINUTE = int(os.getenv("REMINDER_MINUTE", "0"))

EXPENSE_CATS = [
    "🛒 Groceries","☕ Cafe","🏠 Rent","🚌 Transport","📱 Subscriptions","🍽 Dining",
    "👕 Clothing","💊 Health","🎮 Entertainment","🏋 Sports","✈️ Travel",
    "📚 Education","🔧 Home","💰 Investments","📦 Other"
]
INCOME_CATS = ["💼 Salary","🖥 Freelance","📈 Dividends","🎁 Gift","🏠 Rental income","💡 Other income"]
HEADERS = ["Date","Amount","Category","Description","Type","Month","UserID"]

pending = {}

# ─── GOOGLE SHEETS ─────────────────────────────────────────────────────────────
_gs_client = None
_spreadsheet = None

def get_spreadsheet():
    global _gs_client, _spreadsheet
    if _spreadsheet:
        return _spreadsheet
    try:
        import gspread, json
        from google.oauth2.service_account import Credentials
        scopes = ["https://www.googleapis.com/auth/spreadsheets","https://www.googleapis.com/auth/drive"]
        creds_json = os.getenv("GOOGLE_CREDS_JSON")
        if creds_json:
            creds_info = json.loads(creds_json)
            creds = Credentials.from_service_account_info(creds_info, scopes=scopes)
        else:
            logger.warning("No Google credentials found")
            return None
        _gs_client = gspread.authorize(creds)
        _spreadsheet = _gs_client.open(SPREADSHEET_NAME)
        return _spreadsheet
    except Exception as e:
        logger.error(f"Google Sheets error: {e}")
        return None

def get_user_sheet(user_id: int):
    sp = get_spreadsheet()
    if not sp:
        return None
    sheet_name = f"user_{user_id}"
    try:
        sheet = sp.worksheet(sheet_name)
    except Exception:
        try:
            sheet = sp.add_worksheet(title=sheet_name, rows=5000, cols=8)
            sheet.append_row(HEADERS)
        except Exception as e:
            logger.error(f"Create sheet error: {e}")
            return None
    return sheet

def read_user_tx(user_id: int):
    sheet = get_user_sheet(user_id)
    if not sheet:
        return []
    try:
        return sheet.get_all_records()
    except Exception as e:
        logger.error(f"Read error: {e}")
        return []

def write_user_tx(user_id: int, date, amount, category, description, tx_type):
    sheet = get_user_sheet(user_id)
    if not sheet:
        _write_local(user_id, date, amount, category, description, tx_type)
        return
    try:
        sheet.append_row([
            date.strftime("%d.%m.%Y"), amount, category,
            description, tx_type, date.strftime("%Y-%m"), str(user_id)
        ])
    except Exception as e:
        logger.error(f"Write error: {e}")
        _write_local(user_id, date, amount, category, description, tx_type)

# ─── LOCAL FALLBACK ────────────────────────────────────────────────────────────
def _local_path(user_id): return f"/tmp/{user_id}.csv"

def _ensure_local(user_id):
    p = _local_path(user_id)
    if not os.path.exists(p):
        with open(p,"w",newline="",encoding="utf-8") as f:
            csv.writer(f).writerow(HEADERS)

def _write_local(user_id, date, amount, category, description, tx_type):
    _ensure_local(user_id)
    with open(_local_path(user_id),"a",newline="",encoding="utf-8") as f:
        csv.writer(f).writerow([date.strftime("%d.%m.%Y"),amount,category,description,tx_type,date.strftime("%Y-%m"),str(user_id)])

def _read_local(user_id):
    _ensure_local(user_id)
    with open(_local_path(user_id),"r",encoding="utf-8") as f:
        return list(csv.DictReader(f))

def get_tx(user_id):
    rows = read_user_tx(user_id)
    if not rows:
        rows = _read_local(user_id)
    return rows

# ─── STATS ─────────────────────────────────────────────────────────────────────
def get_month_stats(user_id, month):
    rows = get_tx(user_id)
    txs = [r for r in rows if r.get("Month") == month]
    exp = sum(float(r["Amount"]) for r in txs if str(r.get("Type","")).lower() in ("expense","расход"))
    inc = sum(float(r["Amount"]) for r in txs if str(r.get("Type","")).lower() in ("income","доход"))
    by_cat = {}
    for r in txs:
        if str(r.get("Type","")).lower() in ("expense","расход"):
            by_cat[r["Category"]] = by_cat.get(r["Category"],0) + float(r["Amount"])
    return exp, inc, by_cat

# ─── FINN AI ───────────────────────────────────────────────────────────────────
_gemini_model_cache = None

async def get_gemini_model() -> tuple[str, str] | None:
    """Find first available Gemini model that supports generateContent."""
    global _gemini_model_cache
    if _gemini_model_cache:
        return _gemini_model_cache
    import aiohttp
    try:
        url = f"https://generativelanguage.googleapis.com/v1beta/models?key={GEMINI_API_KEY}&pageSize=50"
        async with aiohttp.ClientSession() as session:
            async with session.get(url, timeout=aiohttp.ClientTimeout(total=10)) as resp:
                data = await resp.json()
        models = data.get("models", [])
        # prefer flash models for speed/cost
        prefer = ["flash", "pro"]
        for keyword in prefer:
            for m in models:
                name = m.get("name", "")
                methods = m.get("supportedGenerationMethods", [])
                if keyword in name and "generateContent" in methods:
                    model_id = name.split("/")[-1]
                    logger.info(f"Selected Gemini model: {model_id}")
                    _gemini_model_cache = ("v1beta", model_id)
                    return _gemini_model_cache
    except Exception as e:
        logger.error(f"get_gemini_model error: {e}")
    return None

async def ask_finn(summary: str, question: str) -> str:
    if not GEMINI_API_KEY:
        return "Please add GEMINI_API_KEY to Railway variables."
    try:
        import aiohttp
        model_info = await get_gemini_model()
        if not model_info:
            return "Couldn't connect to Gemini API. Check your API key!"
        models = [model_info]
        for api_ver, model in models:
            url = f"https://generativelanguage.googleapis.com/{api_ver}/models/{model}:generateContent?key={GEMINI_API_KEY}"
            prompt = (
                "You are Finn, a friendly personal finance assistant in a Telegram bot. "
                "Be concise, helpful and supportive. Use emojis. Max 150 words. "
                f"Always respond in English.\n\nUser data:\n{summary}\n\nQuestion: {question}"
            )
            payload = {
                "contents": [{"parts": [{"text": prompt}]}],
                "generationConfig": {"maxOutputTokens": 1024, "temperature": 0.7},
                "safetySettings": [
                    {"category": "HARM_CATEGORY_HARASSMENT", "threshold": "BLOCK_NONE"},
                    {"category": "HARM_CATEGORY_HATE_SPEECH", "threshold": "BLOCK_NONE"},
                    {"category": "HARM_CATEGORY_SEXUALLY_EXPLICIT", "threshold": "BLOCK_NONE"},
                    {"category": "HARM_CATEGORY_DANGEROUS_CONTENT", "threshold": "BLOCK_NONE"}
                ]
            }
            async with aiohttp.ClientSession() as session:
                async with session.post(url, json=payload, timeout=aiohttp.ClientTimeout(total=25)) as resp:
                    data = await resp.json()
                    logger.info(f"Gemini {model} status={resp.status} keys={list(data.keys())}")
                    if "candidates" in data and data["candidates"]:
                        parts = data["candidates"][0]["content"].get("parts", [])
                        # join all non-thought parts (thinking models return thought + response)
                        text = "".join(
                            p.get("text", "") for p in parts if not p.get("thought", False)
                        )
                        if text.strip():
                            return text.strip()
                    if "error" in data:
                        logger.error(f"Gemini {model} error: {data['error']}")
                        # reset cache so next call tries a different model
                        _gemini_model_cache = None
                        continue
                    logger.warning(f"Gemini {model} full response: {str(data)[:500]}")
        return "I couldn't analyze your data right now. Try a simpler question!"
    except Exception as e:
        logger.error(f"Gemini error: {type(e).__name__}: {e}")
        return "I'm having trouble thinking right now. Try again in a moment!"

def build_finn_summary(uid: int) -> str:
    now = datetime.now()
    cur = now.strftime("%Y-%m")
    prev = (now.replace(day=1) - timedelta(days=1)).strftime("%Y-%m")
    exp_c, inc_c, by_cat = get_month_stats(uid, cur)
    exp_p, inc_p, _ = get_month_stats(uid, prev)
    top = sorted(by_cat.items(), key=lambda x: x[1], reverse=True)[:5]
    lines = [
        f"Month: {now.strftime('%B %Y')}",
        f"Expenses: {exp_c:.2f}EUR", f"Income: {inc_c:.2f}EUR",
        f"Balance: {inc_c-exp_c:+.2f}EUR",
        f"Last month expenses: {exp_p:.2f}EUR",
        "Top spending categories:"
    ]
    for cat, amt in top:
        lines.append(f"  {cat}: {amt:.2f}EUR")
    return "\n".join(lines)

async def finn_cmd(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    uid = update.effective_user.id
    name = update.effective_user.first_name or "friend"
    # works for both direct command and button callback
    msg = update.message or (update.callback_query.message if update.callback_query else None)
    if not msg:
        return
    if not ctx.args:
        await msg.reply_text(
            f"Hi {name}! I'm Finn 🦊 your personal finance buddy!\n\n"
            "Ask me anything:\n"
            "`финн how am I doing this month?`\n"
            "`финн where am I overspending?`\n"
            "`финн how can I save more?`\n"
            "`финн compare to last month`\n"
            "`финн give me a saving tip`",
            parse_mode="Markdown"
        )
        return
    question = " ".join(ctx.args)
    await msg.chat.send_action("typing")
    try:
        summary = build_finn_summary(uid)
        response = await ask_finn(summary, question)
        await msg.reply_text(f"🦊 Finn says:\n\n{response}")
    except Exception as e:
        logger.error(f"Finn cmd error: {e}")
        await msg.reply_text("Something went wrong. Try again!")

# ─── HELPERS ──────────────────────────────────────────────────────────────────
def cat_keyboard(tx_type="expense"):
    cats = EXPENSE_CATS if tx_type=="expense" else INCOME_CATS
    buttons, row = [], []
    for cat in cats:
        row.append(InlineKeyboardButton(cat, callback_data=f"cat:{cat}"))
        if len(row)==2: buttons.append(row); row=[]
    if row: buttons.append(row)
    buttons.append([InlineKeyboardButton("❌ Cancel", callback_data="cat:cancel")])
    return InlineKeyboardMarkup(buttons)

def main_keyboard():
    buttons = []
    if WEBAPP_URL:
        buttons.append([InlineKeyboardButton("📊 Open Dashboard", web_app=WebAppInfo(url=WEBAPP_URL))])
    buttons.append([
        InlineKeyboardButton("📋 /stats", callback_data="cmd:stats"),
        InlineKeyboardButton("📈 /compare", callback_data="cmd:compare"),
    ])
    buttons.append([
        InlineKeyboardButton("🦊 /finn", callback_data="cmd:finn"),
        InlineKeyboardButton("❓ /help", callback_data="cmd:help"),
    ])
    return InlineKeyboardMarkup(buttons)

# ─── COMMANDS ─────────────────────────────────────────────────────────────────
async def start(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    uid = update.effective_user.id
    name = update.effective_user.first_name or "there"
    sheet = get_user_sheet(uid)
    storage = "Google Sheets ✅" if sheet else "local storage ⚠️"
    await update.message.reply_text(
        f"👋 Hi {name}! Welcome to *Finance Bot*\n\n"
        f"Storage: {storage}\n\n"
        "📝 *How to add transactions:*\n"
        "`coffee 4.5` — expense\n"
        "`salary +3411` — income\n\n"
        "📋 *Commands:*\n"
        "/stats — monthly breakdown\n"
        "/compare — vs last month\n"
        "/finn — AI finance assistant\n"
        "/settings — notification time\n"
        "/export — download CSV\n"
        "/last — last 10 transactions\n"
        "/deletedata — delete all data",
        parse_mode="Markdown",
        reply_markup=main_keyboard()
    )

async def help_cmd(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    msg = update.message or (update.callback_query.message if update.callback_query else None)
    if msg:
        await msg.reply_text(
            "📖 *Finance Bot Help*\n\n"
            "*Adding transactions:*\n"
            "`coffee 4.5` → expense\n"
            "`salary +3411` → income\n"
            "`coffee 4.5 25.05` → with date\n\n"
            "*Commands:*\n"
            "/stats `[YYYY-MM]` — monthly stats\n"
            "/compare — this vs last month\n"
            "/finn — AI assistant\n"
            "/settings `HH:MM` — set reminder time\n"
            "/export — CSV export\n"
            "/find `query` — search\n"
            "/last — last 10 transactions\n"
            "/budget `category limit` — spending limit\n"
            "/deletedata — delete all data",
            parse_mode="Markdown"
        )

async def stats_cmd(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    uid = update.effective_user.id
    month = ctx.args[0] if ctx.args else datetime.now().strftime("%Y-%m")
    exp, inc, by_cat = get_month_stats(uid, month)
    lines = [f"📊 *Stats for {month}*\n"]
    if by_cat:
        for cat, amt in sorted(by_cat.items(), key=lambda x: x[1], reverse=True):
            bar = "▓" * min(int(amt/50),8)
            lines.append(f"{cat}: *{amt:.2f}€* {bar}")
    else:
        lines.append("_No expenses yet_")
    lines.append(f"\n💸 Total expenses: *{exp:.2f}€*")
    if inc > 0:
        lines.append(f"💚 Total income: *{inc:.2f}€*")
        lines.append(f"📈 Balance: *{inc-exp:.2f}€*")
    msg = update.message or (update.callback_query.message if update.callback_query else None)
    if msg: await msg.reply_text("\n".join(lines), parse_mode="Markdown")

async def compare_cmd(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    uid = update.effective_user.id
    now = datetime.now()
    cur = now.strftime("%Y-%m")
    prev = (now.replace(day=1)-timedelta(days=1)).strftime("%Y-%m")
    exp_c,inc_c,by_c = get_month_stats(uid, cur)
    exp_p,inc_p,by_p = get_month_stats(uid, prev)
    diff = exp_c - exp_p
    sign = "+" if diff>=0 else ""
    pct = (diff/exp_p*100) if exp_p>0 else 0
    lines = [f"📊 *{cur} vs {prev}*\n"]
    lines.append(f"Expenses: *{exp_c:.2f}€* vs {exp_p:.2f}€ ({sign}{diff:.2f}€, {sign}{pct:.0f}%)")
    lines.append(f"Income: *{inc_c:.2f}€* vs {inc_p:.2f}€\n")
    all_cats = set(list(by_c.keys())+list(by_p.keys()))
    for cat in sorted(all_cats, key=lambda c: by_c.get(c,0), reverse=True)[:6]:
        c=by_c.get(cat,0); p=by_p.get(cat,0); d=c-p
        arrow="↑" if d>0 else ("↓" if d<0 else "→")
        lines.append(f"{cat}: {c:.0f}€ {arrow} ({'+' if d>=0 else ''}{d:.0f}€)")
    msg = update.message or (update.callback_query.message if update.callback_query else None)
    if msg: await msg.reply_text("\n".join(lines), parse_mode="Markdown")

async def export_cmd(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    uid = update.effective_user.id
    rows = get_tx(uid)
    msg = update.message or (update.callback_query.message if update.callback_query else None)
    if not rows:
        if msg: await msg.reply_text("No transactions to export yet.")
        return
    output = io.StringIO()
    writer = csv.DictWriter(output, fieldnames=HEADERS)
    writer.writeheader()
    writer.writerows(rows)
    output.seek(0)
    filename = f"my_finances_{datetime.now().strftime('%Y%m')}.csv"
    if msg:
        await msg.reply_document(
            document=io.BytesIO(output.getvalue().encode("utf-8")),
            filename=filename,
            caption=f"📤 Your transactions — {len(rows)} records",
        )

async def find_cmd(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    uid = update.effective_user.id
    if not ctx.args:
        await update.message.reply_text("Usage: `/find coffee`", parse_mode="Markdown")
        return
    query = " ".join(ctx.args).lower()
    rows = get_tx(uid)
    found = [r for r in rows if query in r.get("Description","").lower() or query in r.get("Category","").lower()]
    if not found:
        await update.message.reply_text(f"Nothing found for «{query}»")
        return
    found = list(reversed(found[-10:]))
    lines = [f"🔍 *Results for «{query}»*\n"]
    for r in found:
        sign = "+" if str(r.get("Type","")).lower() in ("income","доход") else "-"
        lines.append(f"{r['Date']} | {sign}{r['Amount']}€ | {r['Category']} | {r['Description']}")
    await update.message.reply_text("\n".join(lines), parse_mode="Markdown")

async def last_cmd(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    uid = update.effective_user.id
    rows = get_tx(uid)
    last = list(reversed(rows[-10:])) if rows else []
    if not last:
        await update.message.reply_text("No transactions yet.")
        return
    lines = ["📋 *Last 10 transactions:*\n"]
    for r in last:
        sign = "+" if str(r.get("Type","")).lower() in ("income","доход") else "-"
        lines.append(f"{r['Date']} | {sign}{r['Amount']}€ | {r['Category']} | {r['Description']}")
    await update.message.reply_text("\n".join(lines), parse_mode="Markdown")

async def budget_cmd(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    uid = update.effective_user.id
    if ctx.args and len(ctx.args)>=2:
        cat_q = " ".join(ctx.args[:-1]).lower()
        try:
            limit = float(ctx.args[-1])
            matched = next((c for c in EXPENSE_CATS if cat_q in c.lower()), " ".join(ctx.args[:-1]))
            write_user_tx(uid, datetime.now(), limit, matched, "__budget__", "budget")
            await update.message.reply_text(f"✅ Budget set: *{matched}* → {limit:.2f}€/month", parse_mode="Markdown")
            return
        except: pass
    month = datetime.now().strftime("%Y-%m")
    _, _, by_cat = get_month_stats(uid, month)
    rows = get_tx(uid)
    budgets = {r["Category"]: float(r["Amount"]) for r in rows if r.get("Type")=="budget"}
    if not budgets:
        await update.message.reply_text("No budgets set.\nUse: `/budget Groceries 400`", parse_mode="Markdown")
        return
    lines = ["🎯 *Monthly Budgets*\n"]
    for cat, lim in budgets.items():
        spent = by_cat.get(cat,0)
        pct = spent/lim*100 if lim>0 else 0
        bar = "█"*int(pct/10)+"░"*(10-int(pct/10))
        status = "🔴" if pct>=100 else ("🟡" if pct>=80 else "🟢")
        lines.append(f"{status} {cat}\n`{bar}` {pct:.0f}%\n{spent:.2f}€ / {lim:.2f}€\n")
    await update.message.reply_text("\n".join(lines), parse_mode="Markdown")

async def settings_cmd(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    uid = update.effective_user.id
    rows = get_tx(uid)
    current = next((r for r in rows if r.get('Type') == 'setting' and r.get('Category') == 'reminder_time'), None)
    current_time = current['Description'] if current else "20:00"
    if ctx.args and len(ctx.args) == 1:
        try:
            parts = ctx.args[0].split(':')
            hour = int(parts[0])
            minute = int(parts[1]) if len(parts) > 1 else 0
            if not (0 <= hour <= 23 and 0 <= minute <= 59):
                raise ValueError
            sheet = get_user_sheet(uid)
            if sheet:
                all_rows = sheet.get_all_values()
                for i, row in enumerate(all_rows[1:], 2):
                    if len(row) > 4 and row[4] == 'setting':
                        sheet.delete_rows(i)
                        break
                sheet.append_row([datetime.now().strftime("%d.%m.%Y"), 0, 'reminder_time', f"{hour:02d}:{minute:02d}", 'setting', datetime.now().strftime("%Y-%m"), str(uid)])
            await update.message.reply_text(
                f"✅ Reminder time set to *{hour:02d}:{minute:02d}*\n\n"
                f"• Daily reminder at {hour:02d}:{minute:02d}\n"
                f"• Weekly summary on Sundays at {hour:02d}:{minute:02d}\n"
                f"• Monthly report on the 1st at 20:00",
                parse_mode="Markdown"
            )
        except:
            await update.message.reply_text("❌ Invalid format. Use: `/settings 19:30`", parse_mode="Markdown")
    else:
        await update.message.reply_text(
            f"⚙️ *Notification Settings*\n\n"
            f"Current reminder time: *{current_time}*\n\n"
            f"To change: `/settings HH:MM`\n"
            f"Example: `/settings 19:30`\n\n"
            f"*Notifications:*\n"
            f"🌙 Daily reminder at your set time\n"
            f"📊 Weekly summary every Sunday\n"
            f"📅 Monthly report on the 1st at 20:00",
            parse_mode="Markdown"
        )

async def deletedata_cmd(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    uid = update.effective_user.id
    keyboard = InlineKeyboardMarkup([[
        InlineKeyboardButton("✅ Yes, delete everything", callback_data="deleteconfirm:yes"),
        InlineKeyboardButton("❌ Cancel", callback_data="deleteconfirm:no"),
    ]])
    await update.message.reply_text(
        "⚠️ *Are you sure?*\n\nThis will permanently delete ALL your transactions and data.",
        parse_mode="Markdown", reply_markup=keyboard
    )

async def exportxls_cmd(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    uid = update.effective_user.id
    rows = get_tx(uid)
    rows = [r for r in rows if str(r.get("Type","")).lower() not in ("budget","template","setting")]
    if not rows:
        await update.message.reply_text("No transactions to export yet.")
        return
    if not XLSX_OK:
        await export_cmd(update, ctx)
        return
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Transactions"
    headers = ["Date", "Amount (€)", "Category", "Description", "Type", "Month"]
    ws.append(headers)
    for row in rows:
        tx_type = str(row.get("Type","expense")).lower()
        amount = float(row.get("Amount",0))
        ws.append([row.get("Date",""), amount if tx_type in ("income","доход") else -amount,
                   row.get("Category",""), row.get("Description",""), tx_type, row.get("Month","")])
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    filename = f"finances_{datetime.now().strftime('%Y%m%d')}.xlsx"
    await update.message.reply_document(document=output, filename=filename,
        caption=f"Your Finance Report: {len(rows)} transactions.")

# ─── MESSAGE HANDLER ──────────────────────────────────────────────────────────
async def handle_message(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    text = update.message.text.strip()
    uid = update.effective_user.id

    # Allow calling Finn without slash: "финн ...", "finn ..."
    lower = text.lower()
    for trigger in ('/finn', 'финн', 'finn'):
        if lower.startswith(trigger):
            question = text[len(trigger):].strip()
            ctx.args = question.split() if question else []
            await finn_cmd(update, ctx)
            return

    date_pattern = r"(\d{1,2}[./]\d{1,2}(?:[./]\d{2,4})?)"
    custom_date = None
    date_match = re.search(date_pattern, text)
    if date_match:
        date_str = date_match.group(1).replace("/",".")
        text_nd = text.replace(date_match.group(1),"").strip()
        try:
            parts = date_str.split(".")
            if len(parts)==2:
                custom_date = datetime.now().replace(day=int(parts[0]),month=int(parts[1]))
            elif len(parts)==3:
                y=int(parts[2]); y=y+2000 if y<100 else y
                custom_date = datetime(y,int(parts[1]),int(parts[0]))
            if custom_date: text=text_nd
        except: custom_date=None
    pattern = r"^([+]?\d+(?:[.,]\d+)?)\s+(.+)$|^(.+?)\s+([+]?\d+(?:[.,]\d+)?)$|^([+]?\d+(?:[.,]\d+)?)$"
    match = re.match(pattern, text, re.IGNORECASE)
    if not match:
        await update.message.reply_text("Didn't understand 🤔 Try: `coffee 4.50` or `salary +3411`", parse_mode="Markdown")
        return
    g = match.groups()
    if g[0] and g[1]: raw,desc=g[0],g[1]
    elif g[2] and g[3]: desc,raw=g[2],g[3]
    else: raw,desc=g[4],"—"
    raw=raw.replace(",",".")
    is_income=raw.startswith("+")
    amount=float(raw.lstrip("+"))
    tx_type="income" if is_income else "expense"
    pending[uid]={"amount":amount,"description":desc,"type":tx_type,"date":custom_date or datetime.now()}
    sign="+" if is_income else "-"
    date_info=f" ({custom_date.strftime('%d.%m.%Y')})" if custom_date else ""
    await update.message.reply_text(
        f"{'💚' if is_income else '💸'} *{sign}{amount:.2f}€* — {desc}{date_info}\n\nChoose category:",
        reply_markup=cat_keyboard(tx_type), parse_mode="Markdown"
    )

# ─── CALLBACK ─────────────────────────────────────────────────────────────────
async def handle_callback(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    uid = query.from_user.id

    if query.data.startswith("cmd:"):
        cmd = query.data.split(":")[1]
        if cmd=="stats": await stats_cmd(update,ctx)
        elif cmd=="compare": await compare_cmd(update,ctx)
        elif cmd=="finn": await finn_cmd(update,ctx)
        elif cmd=="help": await help_cmd(update,ctx)
        return

    if query.data.startswith("deleteconfirm:"):
        ans = query.data.split(":")[1]
        if ans=="yes":
            sheet = get_user_sheet(uid)
            if sheet:
                try:
                    sp = get_spreadsheet()
                    sp.del_worksheet(sheet)
                except: pass
            try:
                os.remove(f"/tmp/{uid}.csv")
            except: pass
            await query.edit_message_text("✅ All your data has been deleted.")
        else:
            await query.edit_message_text("❌ Cancelled. Your data is safe.")
        return

    if query.data=="cat:cancel":
        pending.pop(uid,None)
        await query.edit_message_text("❌ Cancelled.")
        return

    category = query.data.replace("cat:","")
    info = pending.pop(uid,None)
    if not info:
        await query.edit_message_text("Something went wrong, try again.")
        return
    write_user_tx(uid, info["date"], info["amount"], category, info["description"], info["type"])
    sign="+" if info["type"]=="income" else "-"
    await query.edit_message_text(
        f"✅ Saved!\n*{sign}{info['amount']:.2f}€* — {info['description']}\n"
        f"Category: {category}\nDate: {info['date'].strftime('%d.%m.%Y %H:%M')}",
        parse_mode="Markdown"
    )

# ─── SCHEDULER ────────────────────────────────────────────────────────────────
def get_all_user_ids():
    sp = get_spreadsheet()
    if not sp:
        return []
    try:
        return [int(ws.title.replace('user_', '')) for ws in sp.worksheets() if ws.title.startswith('user_')]
    except Exception as e:
        logger.error(f"get_all_user_ids error: {e}")
        return []

async def send_daily_reminder(app):
    user_ids = get_all_user_ids()
    logger.info(f"Daily reminder → {len(user_ids)} users")
    for uid in user_ids:
        try:
            await app.bot.send_message(chat_id=uid,
                text="🌙 *Evening check-in!*\n\nDon't forget to log today's expenses 📝\n\n"
                     "Just send me:\n`coffee 4.5` — expense\n`salary +3000` — income\n\n"
                     "Use /stats to see today's summary.",
                parse_mode="Markdown")
        except Exception as e:
            logger.warning(f"Daily reminder failed for {uid}: {e}")

async def send_weekly_stats(app):
    user_ids = get_all_user_ids()
    now = datetime.now(TIMEZONE)
    for uid in user_ids:
        try:
            rows = get_tx(uid)
            week_txs = []
            for r in rows:
                try:
                    date_str = r.get('Date','')
                    if '.' in date_str:
                        p = date_str.split('.')
                        yr = int(p[2]) if len(p[2])==4 else 2000+int(p[2])
                        tx_date = datetime(yr, int(p[1]), int(p[0]))
                        if tx_date.date() >= (now - timedelta(days=7)).date():
                            week_txs.append(r)
                except: pass
            exp = sum(float(r.get('Amount',0)) for r in week_txs if str(r.get('Type','')).lower() in ('expense','расход'))
            inc = sum(float(r.get('Amount',0)) for r in week_txs if str(r.get('Type','')).lower() in ('income','доход'))
            by_cat = {}
            for r in week_txs:
                if str(r.get('Type','')).lower() in ('expense','расход'):
                    cat = r.get('Category','Other')
                    by_cat[cat] = by_cat.get(cat,0) + float(r.get('Amount',0))
            top = sorted(by_cat.items(), key=lambda x: x[1], reverse=True)[:3]
            lines = [f"📊 *Weekly Summary* ({(now-timedelta(days=7)).strftime('%d.%m')} — {now.strftime('%d.%m')})\n",
                     f"💸 Spent: *{exp:.2f}€*"]
            if inc > 0: lines.append(f"💚 Earned: *{inc:.2f}€*")
            lines.append(f"📈 Net: *{inc-exp:+.2f}€*\n")
            if top:
                lines.append("*Top categories:*")
                for cat, amt in top:
                    lines.append(f"  {cat}: {amt:.2f}€")
            await app.bot.send_message(chat_id=uid, text='\n'.join(lines), parse_mode="Markdown")
        except Exception as e:
            logger.warning(f"Weekly stats failed for {uid}: {e}")

async def send_monthly_stats(app):
    user_ids = get_all_user_ids()
    now = datetime.now(TIMEZONE)
    last_month = (now.replace(day=1) - timedelta(days=1))
    month_key = last_month.strftime("%Y-%m")
    month_name = last_month.strftime("%B %Y")
    for uid in user_ids:
        try:
            exp, inc, by_cat = get_month_stats(uid, month_key)
            if exp == 0 and inc == 0:
                continue
            top = sorted(by_cat.items(), key=lambda x: x[1], reverse=True)[:5]
            lines = [f"📅 *Monthly Report — {month_name}*\n",
                     f"💸 Total spent: *{exp:.2f}€*"]
            if inc > 0:
                lines.append(f"💚 Total earned: *{inc:.2f}€*")
                lines.append(f"📈 Balance: *{inc-exp:+.2f}€*")
            lines.append("")
            if top:
                lines.append("*Spending breakdown:*")
                for cat, amt in top:
                    pct = amt/exp*100 if exp > 0 else 0
                    bar = "▓" * min(int(amt/50), 8)
                    lines.append(f"{cat}: *{amt:.2f}€* ({pct:.0f}%) {bar}")
            lines.append("\nGreat job tracking! Keep it up 💪")
            await app.bot.send_message(chat_id=uid, text='\n'.join(lines), parse_mode="Markdown")
        except Exception as e:
            logger.warning(f"Monthly stats failed for {uid}: {e}")

# ─── MAIN ─────────────────────────────────────────────────────────────────────
def main():
    app = Application.builder().token(TELEGRAM_TOKEN).build()
    app.add_handler(CommandHandler("start", start))
    app.add_handler(CommandHandler("help", help_cmd))
    app.add_handler(CommandHandler("stats", stats_cmd))
    app.add_handler(CommandHandler("compare", compare_cmd))
    app.add_handler(CommandHandler("export", export_cmd))
    app.add_handler(CommandHandler("find", find_cmd))
    app.add_handler(CommandHandler("last", last_cmd))
    app.add_handler(CommandHandler("budget", budget_cmd))
    app.add_handler(CommandHandler("settings", settings_cmd))
    app.add_handler(CommandHandler("deletedata", deletedata_cmd))
    app.add_handler(CommandHandler("exportxls", exportxls_cmd))
    app.add_handler(CommandHandler("finn", finn_cmd))
    app.add_handler(CallbackQueryHandler(handle_callback))
    app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, handle_message))

    # Scheduler — runs inside app event loop
    async def post_init(application):
        try:
            from apscheduler.schedulers.asyncio import AsyncIOScheduler
            from apscheduler.triggers.cron import CronTrigger
            scheduler = AsyncIOScheduler(timezone=TIMEZONE)
            # Daily reminder at default time (users can customize via /settings)
            scheduler.add_job(lambda: asyncio.ensure_future(send_daily_reminder(application)),
                CronTrigger(hour=DEFAULT_REMINDER_HOUR, minute=DEFAULT_REMINDER_MINUTE, timezone=TIMEZONE))
            scheduler.add_job(lambda: asyncio.ensure_future(send_weekly_stats(application)),
                CronTrigger(day_of_week='sun', hour=DEFAULT_REMINDER_HOUR, minute=DEFAULT_REMINDER_MINUTE, timezone=TIMEZONE))
            scheduler.add_job(lambda: asyncio.ensure_future(send_monthly_stats(application)),
                CronTrigger(day=1, hour=20, minute=0, timezone=TIMEZONE))
            scheduler.start()
            logger.info(f"Scheduler started! Daily {DEFAULT_REMINDER_HOUR:02d}:{DEFAULT_REMINDER_MINUTE:02d}, weekly Sun, monthly 1st 20:00")
        except Exception as e:
            logger.error(f"Scheduler error: {e}")

    app.post_init = post_init
    print("🤖 Finance Bot v4 started!")
    app.run_polling(drop_pending_updates=True)

if __name__ == "__main__":
    main()
